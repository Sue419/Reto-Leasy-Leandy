from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.views import View
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from .models import DataTaxi
from .forms import ReportForm
import pandas as pd

class LoginView(View):
    def get(self, request):
        form = AuthenticationForm()
        return render(request, 'login.html', {'form': form})
    
    def post(self, request):
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")
        else:
            messages.error(request, "Usuario y contraseña incorrectos.")
        return render(request, 'login.html', {'form': form})

class DashboardView(LoginRequiredMixin, View):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    def get(self, request):
        data = DataTaxi.objects.all()
        paginator = Paginator(data, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        report_form = ReportForm()
        return render(request, 'dashboard.html', {'page_obj': page_obj, 'report_form': report_form})

    def post(self, request):
        try:
            file = request.FILES['file']
            if file.name.endswith('.csv'):
                self.handle_csv(file)
            elif file.name.endswith('.xlsx'):
                self.handle_xlsx(file)
            else:
                messages.error(request, "Este no es un archivo CSV o XLSX. Por favor, intente de nuevo.")
        except KeyError:
            messages.error(request, "No se ha subido ningún archivo.")
        
        return redirect('dashboard')

    def handle_csv(self, file):
        file_data = file.read().decode('utf-8')
        lines = file_data.split("\n")

        for line in lines:
            fields = line.split(",")
            if len(fields) == 8:  
                DataTaxi.objects.create(
                    first_name=fields[0].strip(),
                    last_name=fields[1].strip(),
                    document_number=fields[2].strip(),
                    contract_start=fields[3].strip(),
                    weekly_fee=fields[4].strip(),
                    car_brand=fields[5].strip(),
                    car_model=fields[6].strip(),
                    car_plate=fields[7].strip()
                )
        messages.success(self.request, "Archivo CSV subido correctamente.")

    def handle_xlsx(self, file):
        wb = load_workbook(file)
        sheet = wb.active
        headers = [cell.value.strip() for cell in sheet[1]]

        if len(headers) != 8 or headers != ['nombres', 'apellidos', 'numero de documento', 'inicio de contrato', 'cuota semanal', 'marca del auto', 'modelo del auto', 'placa del auto']:
            raise ValidationError("El archivo XLSX no tiene las columnas requeridas.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            DataTaxi.objects.create(
                first_name=row[0].strip(),
                last_name=row[1].strip(),
                document_number=row[2].strip(),
                contract_start=row[3].strip(),
                weekly_fee=row[4].strip(),
                car_brand=row[5].strip(),
                car_model=row[6].strip(),
                car_plate=row[7].strip()
            )

class GenerateReportView(View):
    def post(self, request):
        #para generar el reporte Excel...falta más
        messages.success(request, 'Reporte generado exitosamente.')
        return redirect('dashboard')

def logout_view(request):
    logout(request)
    return redirect('login')
